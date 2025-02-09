from openpyxl import load_workbook
import numpy as np

# Membaca data dari file Excel
file_path = "solar_data2.xlsx"
wb = load_workbook(file_path)
sheet = wb.active

# Mengambil nilai dari kolom C (baris 2-97)
values = [sheet.cell(row=i, column=3).value for i in range(2, 98)]

# Menghitung nilai rata-rata
mean_value = np.mean(values)

# Mean Absolute Error (MAE)
mae = np.mean([abs(x - mean_value) for x in values])
mae_percent = (mae / mean_value) * 100

# Mean Squared Error (MSE)
mse = np.mean([(x - mean_value) ** 2 for x in values])

# Standard Deviation (STD)
std = np.std(values)
std_percent = (std / mean_value) * 100

# Menampilkan hasil
print(f"Mean: {mean_value:.8f}")
print(f"Mean Absolute Error (MAE): {mae:.8f} ({mae_percent:.2f}%)")
print(f"Mean Squared Error (MSE): {mse:.8f}")
print(f"Standard Deviation (STD): {std:.8f} ({std_percent:.2f}%)")
